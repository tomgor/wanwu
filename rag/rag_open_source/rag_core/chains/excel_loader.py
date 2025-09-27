import pandas as pd
import logging

import re
import json
from pathlib import Path
from typing import List, Optional, cast
# from langchain.docstore.document import Document
# from langchain.document_loaders import TextLoader
from langchain_core.documents import Document
from langchain_community.document_loaders import TextLoader
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
import sys
import os
import nltk
current_file_path = os.path.abspath(__file__)
# 获取当前文件所在的目录
current_dir = os.path.dirname(current_file_path)
# 添加项目根目录到 sys.path
# sys.path.append(root_dir)
# 拼接nltk_data文件夹的路径
nltk_data_path = os.path.join(current_dir, 'nltk_data')
nltk.data.path.append(nltk_data_path)
# from utils import minio_utils
from logging_config import setup_logging
logger_name = 'rag_excel_loader'
app_name = os.getenv("LOG_FILE")
logger = setup_logging(app_name,logger_name)
logger.info(logger_name+'---------LOG_FILE：'+repr(app_name))
##
# def trans_excel(filepath,sheet_name="故障清单"):
#     data = pd.read_excel(filepath,sheet_name=sheet_name)
#     with open(filepath.replace(".xlsx", ".txt"),"w") as f:
#         for i,d in enumerate(data.values):
#             text = str(d[2]) + "的" + str(d[3]) + "的"+ str(d[4]) + "的"+ str(d[5]) + "的"+ str(d[7]) + "的零部件" +str(d[8])+"出现"+str(d[9])+"。故障现象为" + str(d[14]) + "。根本原因为" +str(d[15])+"。它的解决方案为"+str(d[16])
#             f.write(text+"\n")
#     return filepath.replace(".xlsx", ".txt")
def trans_excel(filepath,del_columns=[]):
    data = pd.read_excel(filepath)
    data.drop_duplicates(inplace=True)
    if del_columns!=[]:
        del data[del_columns]
    columns = data.columns
    columns_len = len(columns)
    with open(filepath.replace(".xlsx", ".txt"),"w") as f:
        for i,d in enumerate(data.values):
            text = ''
            for j in range(columns_len):
                text = text + str(columns[j])+":"+str(d[j])+","
            f.write(text+"\n")
    return filepath.replace(".xlsx", ".txt")

class ExcelLoader(TextLoader):

    def load(self) -> List[Document]:
        """Load from file path."""
        text = ""
        # max_sen_size = 3000
        try:
            path_obj = Path(self.file_path)
            file_name = path_obj.stem
            file_full_name = os.path.split(self.file_path)[-1]
            wb = load_workbook(filename=self.file_path, data_only=True)

            valid_sheet_list = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                total_rows = int(ws.max_row)
                if total_rows > 1:
                    valid_sheet_list.append(sheet_name)
            print(valid_sheet_list)
            for sheet_name in valid_sheet_list:
                ws = wb[sheet_name]
                # ws = wb.active
                # 打印总行数
                total_rows = int(ws.max_row)

                logger.info(f"=======>Total rows in '{file_name}_{sheet_name}': {total_rows}")
                # 获取所有合并单元格的范围
                merged_cells_ranges = ws.merged_cells.ranges
                # 将合并单元格范围转换为更易于处理的格式
                merged_ranges = []
                for rng in merged_cells_ranges:
                    min_row, min_col, max_row, max_col = rng.bounds
                    merged_ranges.append(((min_row, min_col), (max_row, max_col)))
                columns = [cell.value for cell in next(ws.iter_rows()) if cell.value is not None]
                last_column_dict = {}
                if len(valid_sheet_list) > 1:
                    title = "Excel工作簿名称：%s, 每行信息如下:" % sheet_name
                else:
                    title = "Excel中每行信息如下:"
                text += title + "\n"
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=len(columns), values_only=True), start=2):
                    # 检查行是否全为空
                    if all(cell is None for cell in row):
                        continue  # 跳过空行


                    for idx, cell_value in enumerate(row, start=1):

                        # 检查单元格是否在合并单元格范围内
                        column_letter = get_column_letter(idx)  # 直接使用列索引获取列字母
                        # 检查单元格是否在合并单元格范围内
                        cell_position = f'{column_letter}{row_idx}'
                        col_num, row_num = coordinate_from_string(cell_position)
                        col_num_index = column_index_from_string(col_num)  # 将列字母转换为列号
                        column_name = str(columns[col_num_index - 1])
                        cell_coordinate = f"{col_num}{row_num}"
                        cell = ws[cell_coordinate]
                        if cell.hyperlink:
                            get_last_value = f"[{cell_value}]({cell.hyperlink.target})"
                            text = text + column_name + ":" + get_last_value + ";"
                            continue

                        if cell_value is None and any(
                                min_row <= row_num <= max_row and min_col <= col_num_index <= max_col
                                for ((min_col, min_row), (max_col, max_row)) in merged_ranges):
                            # print(f"Cell {cell_position} is part of a merged cell.")
                            get_last_value = str(
                                last_column_dict[column_name]) if column_name in last_column_dict else ""
                            text = text + column_name + ":" + get_last_value + ";"
                            # 根据需求处理合并单元格的值
                        elif len(columns) >= col_num_index:
                            cell_process_value = '' if cell_value is None else str(cell_value).replace("\n", " ")

                            # if '\\' in cell_process_value:
                            #     cell_process_value = eval(repr(cell_process_value.replace('\\\\', '\\')))

                            if column_name:
                                text = text + column_name + ":" + cell_process_value + ";"
                            else:
                                text = text + cell_process_value + ";"
                            # logger.info("========>cell_process_value=%s,text=%s" % (cell_process_value, text))
                            # 若本单元格有值且存在跨单元格合并则缓存到历史dict中
                            if cell_value and any(min_row <= row_num <= max_row and min_col <= col_num_index <= max_col
                                                  for ((min_col, min_row), (max_col, max_row)) in merged_ranges):
                                last_column_dict[column_name] = cell_value

                    # if len(text) > max_sen_size:
                    #     text = text[:max_sen_size]
                    text += "\n"
                    logger.info("=======>row_idx=%s,chunk_size=%s" % (row_idx, len(text)))
            print("=======>len=%s,text=%s" % (len(text),text))

        except ValueError as ve:
            logger.info(f"遇到值错误: {ve}. 请检查文件中的数据是否符合预期格式.")
        except Exception as e:
            raise RuntimeError(f"Error loading {self.file_path}") from e

        metadata = {"source": self.file_path}
        return [Document(page_content=text, metadata=metadata)]



    def load_and_split_doc(self) -> (List[dict]):

        chunks = []
        embedding_content = []
        max_sen_size = 3000
        try:
            path_obj = Path(self.file_path)
            file_name = path_obj.stem
            file_full_name = os.path.split(self.file_path)[-1]
            wb = load_workbook(filename=self.file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = sheet.values
                try:
                    cols = next(data)
                except StopIteration:
                    continue
                df = pd.DataFrame(data, columns=cols)

                df.dropna(how="all", inplace=True)

                for index, row in df.iterrows():
                    page_content = []
                    for col_index, (k, v) in enumerate(row.items()):
                        if pd.notna(v):
                            cell = sheet.cell(
                                row=cast(int, index) + 2, column=col_index + 1
                            )  # +2 to account for header and 1-based index
                            if cell.hyperlink:
                                value = f"[{v}]({cell.hyperlink.target})"
                                page_content.append(f'"{k}":"{value}"')
                            else:
                                page_content.append(f'"{k}":"{v}"')
                    # documents.append(
                    #     Document(page_content=";".join(page_content), metadata={"source": self._file_path})
                    # )
                    text = ";".join(page_content)
                    row_num = cast(int, index) + 2
                    # print("index=%s,text=%s" % (row_num, text))
                    chunk = {"text": text, "type": "text", "embedding_chunks": [],
                             "meta_data": {"chunk_type": "excel", "chunk_size": len(text),
                                           "row_num": row_num, "sheet_name": sheet_name}}
                    chunks.append(chunk)

                    logger.info("=======>row_idx=%s,chunk_size=%s" % (row_num, len(text)))

        except Exception as e:
            import traceback
            logger.error("====> load_and_split_doc error %s" % e)
            logger.error(traceback.format_exc())
            # raise RuntimeError(f"Error loading {self.file_path}") from e

        # metadata = {"source": self.file_path}
        return chunks

    def load_and_split_xls(self) -> (List[dict]):

        chunks = []
        embedding_content = []
        max_sen_size = 3000
        try:

            excel_file = pd.ExcelFile(self.file_path, engine="xlrd")
            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name=sheet_name)
                df.dropna(how="all", inplace=True)

                for index, row in df.iterrows():
                    page_content = []
                    for k, v in row.items():
                        if pd.notna(v):
                            page_content.append(f'"{k}":"{v}"')
                    # documents.append(
                    #     Document(page_content=";".join(page_content), metadata={"source": self._file_path})
                    # )

                    text = ";".join(page_content)
                    row_num = cast(int, index) + 2
                    print("index=%s,text=%s" % (row_num, text))
                    chunk = {"text": text, "type": "text", "embedding_chunks": [],
                             "meta_data": {"chunk_type": "excel", "chunk_size": len(text),
                                           "row_num": row_num, "sheet_name": sheet_name}}
                    chunks.append(chunk)

                    logger.info("=======>row_idx=%s,chunk_size=%s" % (row_num, len(text)))

        except Exception as e:
            import traceback
            logger.error("====> load_and_split_doc error %s" % e)
            logger.error(traceback.format_exc())
            # raise RuntimeError(f"Error loading {self.file_path}") from e

        # metadata = {"source": self.file_path}
        return chunks

if __name__ == "__main__":
    filepath = "a.xls"
    # loader = ExcelLoader(filepath)
    # docs = loader.load()
    loader = ExcelLoader(filepath, autodetect_encoding=True)
    chunk_type = 1
    #textsplitter = ChineseTextSplitter(chunk_type, pdf=False, excel=True, sentence_size=2048)
    chunks = loader.load_and_split_xls()
    print(json.dumps(chunks,ensure_ascii=False))
    # print(json.dumps(sub_chunks,ensure_ascii=False))
    # for item in sub_chunks:
    #     print(json.dumps(item,ensure_ascii=False))

    
