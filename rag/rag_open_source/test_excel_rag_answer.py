import json
import openpyxl
import pandas as pd
import requests
answer_list = []
# 初始化一个空字符串来存储组合后的output  
answer_output = ""  
column_name=['RAG-searchList','RAG-output']


excel_file_path_label =r"RAG-test.xlsx"
excel_file_path_save = r"RAG-test-result.xlsx"

# 打开Excel文件
workbook = openpyxl.load_workbook(excel_file_path_label)
# 选择指定的工作表
sheet = workbook['RAG基线数据集']   

# 读取Excel文件数据
df = pd.read_excel(excel_file_path_label,usecols = ['序号','问题'])
# 提取所需的两列数据
question_list = df['问题'].astype(str).tolist()

# 设置URL  
url = 'http://localhost:10891/rag/knowledge/stream/search'
user_id = "1"
kb_name = ['rag_base_test']
headers = {
            "Content-Type": "application/json",
            "X-uid": user_id
            }


for i,data in enumerate(column_name):
    column_index = 1
    while sheet.cell(row=1, column=column_index).value is not None:
                column_index += 1
    # 写入列标题
    sheet.cell(row=1, column=column_index, value=data)
     #知识库名设置为切分长度，分别为125/250/500
for j,question in enumerate(question_list):
    payload = json.dumps({
            "knowledgeBase":kb_name, 
            "question":question,
            "threshold":0,
            "topK":5,
            "history":[],
            "stream":True,
            "search_field":"emc",
            })
    # print(payload)
    response = requests.request("POST", url, headers=headers, data=payload,verify=False)
    
    
    
    answer_output = ""  
    answer_searchList = None  
  
    for line in response.iter_lines():  
        if line.startswith(b"data: "):  
            # 去除前缀并解码为str  
            json_str = line[6:].decode('utf-8')  
            try:  
                # 解析JSON字符串为Python字典  
                data = json.loads(json_str)  
                # 提取output字段的值，并添加到answer_output中  
                answer_output += data['data']['output']  
                # 提取searchList，如果它是非空的且answer_searchList尚未设置  
                if data['data']['searchList'] and not answer_searchList:  
                    answer_searchList = data['data']['searchList']  
            except (json.JSONDecodeError, KeyError):  
                # 如果解析失败或缺少关键字段，则忽略或记录错误  
                pass  # 或者打印错误信息  
  
    # 输出组合后的output字符串  
    print(answer_output)  
  
    # 如果找到了非空的searchList，则打印它（在这个例子中它将是None）  
    if answer_searchList:  
        print(answer_searchList)  
    else:  
        print("没有找到非空的searchList")
    
    answer_list.clear()
    answer_list.append(answer_searchList)  # 将结果添加到列表中  
    answer_list.append(answer_output)  # 将结果添加到列表中  
    for i,data in enumerate(answer_list):
        # 确定新列的索引:标题行左起第一个为空的列
        column_index = 1
        while sheet.cell(row=j+2, column=column_index).value is not None:
            column_index += 1
        # 写入数据`
        json_data = json.dumps(data, ensure_ascii=False)  
        sheet.cell(row=j+2, column=column_index, value=json_data)  
   
        # 保存修改后的Excel文件
        workbook.save(excel_file_path_save)
        workbook.close()     
print("全部测试完成，请查看结果")
